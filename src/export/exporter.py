"""Export module for HUD housing data."""

import csv
import json
import sqlite3
from pathlib import Path

EXPORT_DIR = Path(__file__).parent.parent.parent / "data" / "exports"

TABLES = [
    "lihtc_projects",
    "reac_inspections",
    "section8_contracts",
    "multifamily_properties",
    "cross_links",
    "owners",
]


def export_all(conn: sqlite3.Connection, fmt: str = "all") -> dict:
    """Export all tables in the specified format."""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    results = {}

    if fmt in ("csv", "all"):
        results["csv"] = _export_csv(conn)
    if fmt in ("json", "all"):
        results["json"] = _export_json(conn)
    if fmt in ("excel", "all"):
        results["excel"] = _export_excel(conn)
    if fmt in ("markdown", "all"):
        results["markdown"] = _export_markdown(conn)

    return results


def _export_csv(conn: sqlite3.Connection) -> list[str]:
    """Export each table as CSV."""
    files = []
    for table in TABLES:
        rows = conn.execute(f"SELECT * FROM {table}").fetchall()
        if not rows:
            continue

        cols = [desc[0] for desc in conn.execute(f"SELECT * FROM {table} LIMIT 1").description]
        path = EXPORT_DIR / f"{table}.csv"

        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(cols)
            for row in rows:
                writer.writerow(row)

        files.append(str(path))
    return files


def _export_json(conn: sqlite3.Connection) -> str:
    """Export all data as a single JSON file."""
    data = {"metadata": {"project": "HUD Affordable Housing Compliance Tracker"}}

    for table in TABLES:
        rows = conn.execute(f"SELECT * FROM {table}").fetchall()
        cols = [desc[0] for desc in conn.execute(f"SELECT * FROM {table} LIMIT 1").description]
        data[table] = [dict(zip(cols, row)) for row in rows]

    # Add stats
    from src.storage.database import get_stats
    data["stats"] = get_stats(conn)

    path = EXPORT_DIR / "hud_housing.json"
    path.write_text(json.dumps(data, indent=2, default=str))
    return str(path)


def _export_excel(conn: sqlite3.Connection) -> str:
    """Export to styled Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return ""

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0984E3", end_color="0984E3", fill_type="solid")

    first = True
    for table in TABLES:
        rows = conn.execute(f"SELECT * FROM {table}").fetchall()
        if not rows:
            continue

        cols = [desc[0] for desc in conn.execute(f"SELECT * FROM {table} LIMIT 1").description]

        if first:
            ws = wb.active
            ws.title = table
            first = False
        else:
            ws = wb.create_sheet(title=table[:31])  # Excel limits sheet names to 31 chars

        # Headers
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width columns (approximate)
        for col_idx, col_name in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                min(max(len(col_name), 10), 40)
            )

    path = EXPORT_DIR / "hud_housing.xlsx"
    wb.save(str(path))
    return str(path)


def _export_markdown(conn: sqlite3.Connection) -> str:
    """Export summary statistics as Markdown."""
    from src.storage.database import get_stats
    stats = get_stats(conn)

    lines = [
        "# HUD Affordable Housing Compliance Tracker — Data Summary",
        "",
        "## Record Counts",
        "",
        "| Table | Records |",
        "|-------|---------|",
    ]

    for table in TABLES:
        lines.append(f"| {table} | {stats.get(table, 0):,} |")

    lines.extend([
        "",
        "## Key Metrics",
        "",
        f"- **LIHTC States:** {stats.get('lihtc_states', 0)}",
        f"- **Total LIHTC Units:** {stats.get('total_lihtc_units', 0):,}",
        f"- **Avg REAC Inspection Score:** {stats.get('avg_inspection_score', 0)}",
        f"- **Failing Inspections (< 60):** {stats.get('failing_inspections', 0):,}",
        "",
        "## Quality Averages",
        "",
        "| Table | Avg Quality |",
        "|-------|-------------|",
    ])

    for table in ["lihtc_projects", "reac_inspections", "section8_contracts", "multifamily_properties"]:
        q = stats.get(f"{table}_quality_avg", 0.0)
        lines.append(f"| {table} | {q:.3f} |")

    lines.extend([
        "",
        "---",
        "*Built by Nathan Goldberg — nathanmauricegoldberg@gmail.com*",
    ])

    path = EXPORT_DIR / "summary.md"
    path.write_text("\n".join(lines))
    return str(path)
